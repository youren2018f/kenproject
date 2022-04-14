import streamlit as st
import pandas as pd
import openpyxl
from PIL import Image


from st_aggrid import AgGrid
import streamlit as st
import pandas as pd 
import numpy as np
import openpyxl


# data= pd.read_csv('df_sample_data.csv', index_col=0) 
k_file = r"D:\youren\kenproject\110設備維修統計表.xlsx"
wb = openpyxl.load_workbook(k_file)



with st.sidebar:
    add_selectbox = st.sidebar.selectbox(
        "110年設備維修統計",
        wb.sheetnames
    )


# # Using object notation
# add_selectbox = st.sidebar.selectbox(
#     "110年設備維修統計",
#     wb.sheetnames
# )







df = pd.read_excel(k_file,
                     sheet_name="發電機維修紀錄",header=1)

df= df[df["台站"]== "臺東分臺"]
df.index = np.arange(1, len(df) + 1)
st.table(df.fillna(""))





